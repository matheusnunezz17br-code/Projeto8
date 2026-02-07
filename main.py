import openpyxl

book=openpyxl.Workbook()

print(book.sheetnames)

book.create_sheet('Teste')

teste_page = book["Teste"]

teste_page.append(["Produtos", "Quantidade", "Preços"])

teste_page.append(["Computador", 100, "R$4000"])

teste_page.append(["Teclado", 120, "R$100"])

teste_page.append(["Mouse", 9, "R$40"])

book.save("Planilha de Produtos.xlsx")